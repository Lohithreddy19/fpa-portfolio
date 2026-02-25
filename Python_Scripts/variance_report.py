import pandas as pd
import openpyxl
import numpy as np 
from openpyxl.styles import PatternFill, Font

# Load data
budget = pd.read_csv('Raw_Data/budget_2024.csv')
actuals = pd.read_csv('Raw_Data/actuals_2024.csv')

# merge and calculate variance
df = budget.merge(actuals, on= ['month','department','category'])
df['variance_$'] = df['actual_amount'] - df['budget_amount']
df['variance_%'] = df['variance_$'] / df['budget_amount'].replace(0, np.nan)

# Auto generate commentary
def commentary(row):
    pct = row['variance_%']
    dept = row['department']
    cat  = row['category']
    amt  = abs(row['variance_$'])
    direction = 'over' if pct > 0 else 'under'
    if abs(pct) < 0.05:
        return 'Within budget tolerance.'
    elif abs(pct) < 0.15:
        return f'{dept} {cat} {direction} budget by ${amt:,.0f} ({abs(pct):.1%}). Monitor.'
    else:
        return f'MATERIAL: {dept} {cat} {direction} budget by ${amt:,.0f} ({abs(pct):.1%}). Requires explanation.'

df['commentary'] = df.apply(commentary, axis=1)

# Write to Excel with RAG color coding
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Variance Report'

headers = ['Month','Department','Category','Budget','Actuals','Variance $','Variance %','Commentary']
ws.append(headers)

red    = PatternFill('solid', fgColor='FFB3B3')
amber  = PatternFill('solid', fgColor='FFE5B3')
green  = PatternFill('solid', fgColor='C6F4C2')

for _, row in df.iterrows():
    ws.append([row['month'], row['department'], row['category'],
               row['budget_amount'], row['actual_amount'],
               row['variance_$'], round(row['variance_%'],4), row['commentary']])
    last_row = ws.max_row
    pct = abs(row['variance_%'])
    fill = red if pct > 0.15 else (amber if pct > 0.07 else green)
    for col in range(1, 9):
        ws.cell(last_row, col).fill = fill

wb.save('Excel_Models/Variance_Report_Auto.xlsx')
print('Variance report generated!')

